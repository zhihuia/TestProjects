#Author: Cheese
#Time: 2018/5/28 21:14
from openpyxl import load_workbook

class DoExcel:
    '''
    专门用来与excel的数据文件进行交互
    '''

    def __init__(self,filepath):
        self.wb = load_workbook(filepath)
        self.sh = self.wb["case_datas"]
        self.init_data_sh = self.wb["init_data"]


    #读取所有测试数据
    def get_all_caseDatas(self):
        #先获取所有的初始化数据
        init_datas = self.get_init_data()
        # 定义一个列表，专门用来存储所有的测试用例。每一个测试用例都是一个字典，有所有的相关数据
        all_caseDatas = []
        for row in range(2,self.sh.max_row+1):
            new_case = {}
            new_case["method"] = self.sh.cell(row,column=5).value
            new_case["url"] = self.sh.cell(row, column=6).value
            new_case["expected_data"] = self.sh.cell(row, column=8).value
            #new_case["request_data"] = self.sh.cell(row, column=7).value
            #request_data先不赋值，拿到excel中的请求数据，用一个临时变量接收
            temp_request_data = self.sh.cell(row, column=7).value
            #遍历所有的初始化值的键名，如果请求数据中，有某一个键名，则直接替换
            if temp_request_data is not None:
                for item,value in init_datas.items():     #字典的遍历
                    # 判断一下本条测试数据中，请求数据是否有需要替换的
                    if temp_request_data.find(item) != -1:    #!=-1说明存在，find函数
                        temp_request_data = temp_request_data.replace(item,str(value))
                #替换请求数据
                new_case["request_data"] = temp_request_data
            all_caseDatas.append(new_case)
        return all_caseDatas

    def get_init_data(self):
        init_datas = {}
        #读取所有的初始化数据，第一列为键名，第二列为键值
        for row in range(2,self.init_data_sh.max_row+1):
            key = self.init_data_sh.cell(row,column=1).value
            value = self.init_data_sh.cell(row,column=2).value
            init_datas[key] = value
        #对初始化手机号码进行递增处理，需要2个
        phone2 = int(init_datas["${phone1}"]) + 1
        init_datas["${phone2}"] = str(phone2)
        init_datas["${phone3}"] = str(phone2 + 2)
        #print("init_datas",init_datas)
        return init_datas

    def update_init_phone(self):
        init_phone = self.init_data_sh.cell(2,2).value
        self.init_data_sh.cell(2,2).value = str((int(init_phone)) + 4)

    def save_data(self,filepath):
        self.wb.save(filepath)



















#de = DoExcel("D:\\PythonStudy\\python6_basic\\API\\TestDatas\\api_info_1.xlsx")
#de = DoExcel("/Users/gemii/TestProjects/API_Framwork/TestDatas/api_info_1.xlsx")
#de.get_all_caseDatas()