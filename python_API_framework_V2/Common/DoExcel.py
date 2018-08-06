from openpyxl import load_workbook
import logging

class DoExcel:

    def __init__(self,excelpath):
        self.wb = load_workbook(excelpath)
        self.sh_case_data = self.wb["case_datas"]
        self.sh_prepare_data = self.wb["prepare_datas"]

    #从excel当中读取初始化数据，并对手机号码进行处理。
    def get_init_datas(self):
        init_datas = {}
        for index in range(2, self.sh_prepare_data.max_row + 1):
            key = self.sh_prepare_data.cell(row=index,column=1).value
            init_datas[key] = str(self.sh_prepare_data.cell(row=index,column=2).value)

        init_datas["${phone}"] = str(int(init_datas["${init_phone}"]) + 1)
        init_datas["${noReg_phone}"] = str(int(init_datas["${init_phone}"]) + 2)
        logging.info("初始化数据为：{0}".format(init_datas))
        #{"${init_phone}":"18612340000","${phone}":"18612340001","${phone}":"18612340002"}
        return init_datas

    def get_caseDatas_all(self):
        all_case_datas = []

        for index in range(2,self.sh_case_data.max_row+1):
            case_data = {}
            case_data["case_id"] = self.sh_case_data.cell(row=index,column=1).value
            case_data["api_name"] = self.sh_case_data.cell(row=index, column=4).value
            case_data["method"] = self.sh_case_data.cell(row=index, column=5).value
            case_data["url"] = self.sh_case_data.cell(row=index, column=6).value
            # 比对类型
            case_data["compare_type"] = self.sh_case_data.cell(row=index, column=9).value

            # 获取初始值
            init_datas = self.get_init_datas()

            request_data = self.sh_case_data.cell(row=index, column=7).value
            expected_data = self.sh_case_data.cell(row=index, column=8).value
            # 遍历所有的初始化值的键名。如果请求数据当中，有某一个键名，则直接替换。
            case_data["request_data"] = self._replace_action(init_datas,request_data)
            case_data["expected_data"] = self._replace_action(init_datas,expected_data)

            #判断本用例是否需要对响应结果进行解析，并获取其中的值。
            expression = self.sh_case_data.cell(row=index, column=10).value
            if expression is not None:
                case_data["related_exp"] = self._replace_action(init_datas,expression)

            #将处理后的测试用例数据添加到列表中
            all_case_datas.append(case_data)
        return all_case_datas

    #更新初始化数据
    def update_init_data(self):
        init_data = self.sh_prepare_data.cell(row=2,column=2).value
        self.sh_prepare_data.cell(row=2, column=2).value = str(int(init_data) + 3)

    #保存数据
    def save_excelFile(self,excelpath):
        try:
            self.wb.save(excelpath)
        except Exception as e:
            logging.exception(e)

    #替换操作：
    def _replace_action(self,datas_dict,to_be_replace_str):
        '''
        遍历字典当中的key，如果在被替换的字符串当中，找到了某一个key，则用对应的value替换key。
        在遍历操作结束之后，返回替换后的字符串。
        若没有找到任何一个key，则返回原字符串。
        :param datas_dict: 字典数据。
        :param to_be_replace_str: 要被替换的字符串
        :return: 返回替换之后的字符串，否则返回原字符串
        '''
        if to_be_replace_str is not None and len(datas_dict) > 0:
            for item, value in datas_dict.items():
               if to_be_replace_str.find(item) != -1:
                   to_be_replace_str = to_be_replace_str.replace(item, value)
        return to_be_replace_str

